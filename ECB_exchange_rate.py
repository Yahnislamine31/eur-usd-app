import streamlit as st
import requests
import pandas as pd
from io import StringIO, BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import tempfile
import time
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from decimal import Decimal, ROUND_HALF_UP

# ── PAGE CONFIG ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Ancillary data downloader",
    page_icon="🌐",
    layout="centered",
)

# ── LIGHT THEME CSS ───────────────────────────────────────────────────────────
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600&family=IBM+Plex+Mono:wght@400;600&display=swap');

    html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
    .stApp { background-color: #f8f9fb; color: #1a1d23; }

    .app-header {
        background: linear-gradient(135deg, #003366 0%, #0055a4 100%);
        border-radius: 10px;
        padding: 1.4rem 1.8rem 1.2rem;
        margin-bottom: 1.5rem;
    }
    .app-header h1 {
        font-family: 'IBM Plex Mono', monospace;
        color: #ffffff;
        font-size: 1.6rem;
        font-weight: 600;
        margin: 0 0 0.3rem 0;
    }
    .app-header p { color: #a8c8f0; font-size: 0.88rem; margin: 0; }

    .info-banner {
        background: #eef4ff;
        border: 1px solid #b8d0f5;
        border-left: 4px solid #0055a4;
        border-radius: 6px;
        padding: 0.55rem 1rem;
        font-size: 0.84rem;
        color: #1a3a6b;
        margin-bottom: 0.9rem;
    }
    .warn-banner {
        background: #fff8e6;
        border: 1px solid #f5d580;
        border-left: 4px solid #e6a817;
        border-radius: 6px;
        padding: 0.5rem 1rem;
        font-size: 0.83rem;
        color: #7a5000;
        margin-top: 0.4rem;
    }
    .stButton > button {
        background: #003366 !important;
        color: #fff !important;
        font-family: 'Inter', sans-serif !important;
        font-weight: 500 !important;
        border: none !important;
        border-radius: 5px !important;
        padding: 0.35rem 0.5rem !important;
        width: 100% !important;
        font-size: 0.80rem !important;
        white-space: nowrap !important;
        min-height: 0 !important;
        line-height: 1.3 !important;
    }
    .stButton > button:hover { background: #0055a4 !important; }
    .stButton > button[kind="primary"] {
        font-family: 'IBM Plex Mono', monospace !important;
        font-weight: 600 !important;
        border-radius: 7px !important;
        padding: 0.65rem 2rem !important;
        font-size: 0.98rem !important;
        white-space: normal !important;
        min-height: auto !important;
        line-height: normal !important;
    }
    .stDownloadButton > button {
        background: #1a7f37 !important;
        color: #fff !important;
        font-family: 'IBM Plex Mono', monospace !important;
        font-weight: 600 !important;
        border: none !important;
        border-radius: 7px !important;
        width: 100% !important;
    }
    .stCheckbox label { font-size: 0.94rem; color: #2d3748; }
    div[data-testid="stExpander"] {
        background: #ffffff !important;
        border: 1px solid #e2e6ed !important;
        border-radius: 8px;
        box-shadow: 0 1px 3px rgba(0,0,0,0.05);
        margin-bottom: 0.8rem;
    }
    hr { border-color: #e2e6ed !important; }
</style>
""", unsafe_allow_html=True)


# ── CURRENCY CATALOGUE ────────────────────────────────────────────────────────
# All currencies available from ECB SDW (quoted against EUR)
# Format: code → (full name, region/flag hint)
CURRENCY_CATALOGUE: dict[str, str] = {
    "EUR": "Euro",
    "USD": "US Dollar",
    "GBP": "British Pound",
    "JPY": "Japanese Yen",
    "CHF": "Swiss Franc",
    "CAD": "Canadian Dollar",
    "AUD": "Australian Dollar",
    "NZD": "New Zealand Dollar",
    "NOK": "Norwegian Krone",
    "SEK": "Swedish Krona",
    "DKK": "Danish Krone",
    "SGD": "Singapore Dollar",
    "HKD": "Hong Kong Dollar",
    "CNY": "Chinese Yuan",
    "KRW": "South Korean Won",
    "INR": "Indian Rupee",
    "BRL": "Brazilian Real",
    "MXN": "Mexican Peso",
    "ZAR": "South African Rand",
    "TRY": "Turkish Lira",
    "PLN": "Polish Zloty",
    "CZK": "Czech Koruna",
    "HUF": "Hungarian Forint",
    "RON": "Romanian Leu",
    "BGN": "Bulgarian Lev",
    "HRK": "Croatian Kuna",
    "RUB": "Russian Ruble",
    "IDR": "Indonesian Rupiah",
    "MYR": "Malaysian Ringgit",
    "PHP": "Philippine Peso",
    "THB": "Thai Baht",
    "ILS": "Israeli Shekel",
}

# Currency options for the multiselect: "USD — US Dollar"
CURRENCY_OPTIONS = [f"{code} — {name}" for code, name in CURRENCY_CATALOGUE.items()]


# ── WORLD BANK CONSTANTS ──────────────────────────────────────────────────────
WORLD_BANK_BASE = "https://api.worldbank.org/v2"
HEADER_COLOR    = "003366"
ALTERNATE_ROW   = "EEF3FA"

WB_INDICATOR_GROUPS: dict[str, dict[str, str]] = {
    "📈 Growth & Output": {
        "GDP growth (annual %)":           "NY.GDP.MKTP.KD.ZG",
        "GDP, constant prices (USD)":      "NY.GDP.MKTP.KD",
        "GDP per capita, constant prices": "NY.GDP.PCAP.KD",
        "GDP deflator (annual %)":         "NY.GDP.DEFL.KD.ZG",
    },
    "👥 Population": {
        "Population, total":               "SP.POP.TOTL",
        "Population growth (annual %)":    "SP.POP.GROW",
    },
    "💼 Labour Market": {
        "Unemployment rate (%)":              "SL.UEM.TOTL.ZS",
        "Labor force, total":                 "SL.TLF.TOTL.IN",
        "Labor force participation rate (%)": "SL.TLF.CACT.ZS",
        "GDP per person employed (USD)":      "SL.GDP.PCAP.EM.KD",
    },
    "💰 Prices & Savings": {
        "Inflation, consumer prices (%)":     "FP.CPI.TOTL.ZG",
        "Gross capital formation (% of GDP)": "NE.GDI.TOTL.ZS",
        "Gross savings (% of GDP)":           "NY.GNS.ICTR.ZS",
    },
    "🌍 Trade & External": {
        "Current account balance (% of GDP)":  "BN.CAB.XOKA.GD.ZS",
        "Exports of goods & services (% GDP)": "NE.EXP.GNFS.ZS",
        "Imports of goods & services (% GDP)": "NE.IMP.GNFS.ZS",
        "FDI, net inflows (% of GDP)":         "BX.KLT.DINV.WD.GD.ZS",
    },
    "🏛️ Fiscal": {
        "General government debt (% of GDP)":  "GC.DOD.TOTL.GD.ZS",
        "Central govt revenue (% of GDP)":     "GC.REV.XGRT.GD.ZS",
        "Central govt expenditure (% of GDP)": "GC.XPN.TOTL.GD.ZS",
    },
}

WB_INDICATORS: dict[str, str] = {
    label: code
    for grp in WB_INDICATOR_GROUPS.values()
    for label, code in grp.items()
}

WB_INDICATOR_NOTES: dict[str, str] = {
    "NY.GDP.MKTP.KD.ZG":    "Constant 2015 USD. Annual % change.",
    "NY.GDP.MKTP.KD":       "Constant 2015 USD. Not inflation-adjusted.",
    "NY.GDP.PCAP.KD":       "Constant 2015 USD per capita.",
    "NY.GDP.DEFL.KD.ZG":    "Annual % change in implicit price deflator.",
    "SP.POP.TOTL":          "De facto population, mid-year estimates.",
    "SP.POP.GROW":          "Annual population growth rate (%).",
    "SL.UEM.TOTL.ZS":       "ILO modelled estimates. % of total labour force.",
    "SL.TLF.TOTL.IN":       "Total labour force (persons).",
    "SL.TLF.CACT.ZS":       "Labour force as % of population ages 15+.",
    "SL.GDP.PCAP.EM.KD":    "Constant 1990 PPP USD per employed person.",
    "FP.CPI.TOTL.ZG":       "Consumer price index, annual % change.",
    "NE.GDI.TOTL.ZS":       "Gross capital formation as % of GDP.",
    "NY.GNS.ICTR.ZS":       "Gross savings as % of GDP.",
    "BN.CAB.XOKA.GD.ZS":    "Current account balance as % of GDP.",
    "NE.EXP.GNFS.ZS":       "Exports of goods and services as % of GDP.",
    "NE.IMP.GNFS.ZS":       "Imports of goods and services as % of GDP.",
    "BX.KLT.DINV.WD.GD.ZS": "FDI net inflows as % of GDP.",
    "GC.DOD.TOTL.GD.ZS":    "Central + sub-national govt debt as % of GDP.",
    "GC.REV.XGRT.GD.ZS":    "General government revenue as % of GDP.",
    "GC.XPN.TOTL.GD.ZS":    "General government expenditure as % of GDP.",
}

ECB_SOURCE = {
    "source_name": "European Central Bank (ECB)",
    "source_url":  "https://data-api.ecb.europa.eu/service/data/EXR/",
    "notes":       "Statistical Data Warehouse — EXR series, daily (business days) reference rates. "
                   "Cross rates derived via EUR. Weekly/Monthly/Quarterly/Annual figures are simple "
                   "averages of the daily rate, computed transparently in the generated Stata .do file.",
}


# ── HELPERS ───────────────────────────────────────────────────────────────────
def precise_round(value, precision=4):
    if pd.isna(value):
        return value
    return float(Decimal(str(value)).quantize(
        Decimal("1." + "0" * precision), rounding=ROUND_HALF_UP
    ))


def write_headers(ws, headers, row=1):
    hfill = PatternFill("solid", fgColor=HEADER_COLOR)
    hfont = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center")


def style_row(ws, r, ncols, even):
    if even:
        for c in range(1, ncols + 1):
            ws.cell(r, c).fill = PatternFill("solid", fgColor=ALTERNATE_ROW)


def set_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def col_letter(n: int) -> str:
    """1-based column index → Excel letter (1→A, 27→AA)."""
    result = ""
    while n:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


# ── ECB FETCHER ───────────────────────────────────────────────────────────────
def fetch_ecb_vs_eur(
    currencies: list[str],
    start_date,
    end_date,
) -> pd.DataFrame:
    """
    Fetch all selected currencies against EUR in ONE API call (daily, business-day series).
    ECB URL: EXR/D.{CUR1}+{CUR2}+...EUR.SP00.A

    Returns a wide DataFrame indexed by date with one column per currency
    (values = units of that currency per 1 EUR).
    EUR itself is added as a constant 1.0 column so cross rates work uniformly.

    Only the daily series is ever fetched here — Weekly/Monthly/Quarterly/Annual
    figures are derived later, transparently, in the generated Stata .do file.
    """
    cur_str = "+".join(currencies)
    url = f"https://data-api.ecb.europa.eu/service/data/EXR/D.{cur_str}.EUR.SP00.A"
    params = {
        "startPeriod": start_date.strftime("%Y-%m-%d"),
        "endPeriod":   end_date.strftime("%Y-%m-%d"),
        "format":      "csvdata",
    }

    raw  = pd.read_csv(StringIO(requests.get(url, params=params, timeout=100).text))
    # ECB CSV has CURRENCY and TIME_PERIOD columns
    df   = raw[["CURRENCY", "TIME_PERIOD", "OBS_VALUE"]].copy()
    df.columns = ["Currency", "Date", "Rate"]
    df["Rate"] = pd.to_numeric(df["Rate"], errors="coerce")
    df = df.dropna()

    # Pivot to wide: rows = date, columns = currency code
    wide = df.pivot(index="Date", columns="Currency", values="Rate")
    wide.index.name = "Date"

    # Add EUR = 1.0 so EUR/XXX cross rates work without special-casing
    wide["EUR"] = 1.0

    # Keep only requested currencies (+ EUR) in consistent order
    keep = ["EUR"] + [c for c in currencies if c in wide.columns]
    wide = wide[keep].sort_index()
    return wide


# ── EXCEL BUILDERS ────────────────────────────────────────────────────────────
def build_fx_daily_sheet(wb: Workbook, base: str, quote: str, start_date, end_date):
    """
    Build the single "ECB - FX - Daily" sheet for one currency pair.
    Columns: Date | BASE/QUOTE | QUOTE/BASE  (business days only, as published by the ECB).

    This is the ONLY FX sheet the Excel file contains. Weekly/Monthly/Quarterly/Annual
    figures are not pre-computed here — the generated Stata .do file derives them from
    this daily series (simple average of the daily rate within each period), so the
    calculation stays fully visible rather than being baked into the spreadsheet.
    """
    # Currencies needed — if one is EUR we only need the other; ECB always quotes vs EUR
    non_eur = [c for c in [base, quote] if c != "EUR"]
    wide = fetch_ecb_vs_eur(non_eur, start_date, end_date)

    wide.index = pd.to_datetime(wide.index)
    wide = wide.sort_index()
    wide.index = wide.index.strftime('%Y-%m-%d')
    wide.index.name = "Date"

    fwd_col = f"{base}/{quote}"       # e.g. USD/EUR
    inv_col = f"{quote}/{base}"       # e.g. EUR/USD

    b = wide[base].values   # numpy array — no ambiguity
    q = wide[quote].values
    pairs = pd.DataFrame(index=wide.index)
    pairs[fwd_col] = pd.Series(q / b, index=wide.index).apply(precise_round)
    pairs[inv_col] = pd.Series(b / q, index=wide.index).apply(precise_round)

    ws = wb.active
    ws.title = "ECB - FX - Daily"
    write_headers(ws, ["Date", fwd_col, inv_col])
    set_widths(ws, {"A": 14, "B": 16, "C": 16})
    for i, (date, row) in enumerate(pairs.iterrows()):
        r = i + 2
        ws.cell(r, 1, date)
        ws.cell(r, 2, row[fwd_col])
        ws.cell(r, 3, row[inv_col])
        style_row(ws, r, 3, i % 2 == 0)


# ── WORLD BANK FETCHER ────────────────────────────────────────────────────────
@st.cache_data(show_spinner=False, ttl=3600)
def fetch_wb_indicator(
    indicator_code: str,
    countries: tuple,
    start_year: int,
    end_year: int,
    max_retries: int = 4,
) -> pd.DataFrame:
    fetch_all   = list(countries) == ["all"]
    country_str = "all" if fetch_all else ";".join(countries)
    url         = f"{WORLD_BANK_BASE}/country/{country_str}/indicator/{indicator_code}"
    page, per_page = 1, 1000
    all_rows: list[dict] = []

    while True:
        params = {
            "date":     f"{start_year}:{end_year}",
            "format":   "json",
            "per_page": per_page,
            "page":     page,
        }
        last_exc = None
        for attempt in range(max_retries):
            try:
                resp = requests.get(url, params=params, timeout=60)
                if resp.status_code in (502, 503, 504):
                    raise requests.HTTPError(f"HTTP {resp.status_code}")
                resp.raise_for_status()
                last_exc = None
                break
            except (requests.RequestException, requests.HTTPError) as exc:
                last_exc = exc
                time.sleep(2 ** attempt)

        if last_exc:
            raise last_exc

        data = resp.json()
        if not isinstance(data, list) or len(data) < 2 or not data[1]:
            break

        for item in data[1]:
            country_id = item.get("country", {}).get("id", "")
            if fetch_all and len(country_id) != 2:
                continue
            all_rows.append({
                "Country":      item.get("country", {}).get("value", ""),
                "Country Code": item.get("countryiso3code") or country_id,
                "Year":         int(item["date"]),
                "Value":        item["value"],
            })

        total_pages = data[0].get("pages", 1)
        if page >= total_pages:
            break
        page += 1

    if not all_rows:
        return pd.DataFrame()

    df = pd.DataFrame(all_rows)
    df["Value"] = pd.to_numeric(df["Value"], errors="coerce")
    df = df.dropna(subset=["Value"])
    df = df.sort_values(["Country", "Year"]).reset_index(drop=True)
    return df


@st.cache_data(show_spinner=False, ttl=86400)
def get_wb_countries() -> pd.DataFrame:
    url    = f"{WORLD_BANK_BASE}/country"
    params = {"format": "json", "per_page": 500}
    resp   = requests.get(url, params=params, timeout=20)
    resp.raise_for_status()
    data   = resp.json()
    rows   = [
        {"name": c["name"], "iso2": c["id"]}
        for c in data[1]
        if c.get("region", {}).get("id") != "NA"
    ]
    return pd.DataFrame(rows).sort_values("name").reset_index(drop=True)


def _build_wb_sheet_from_df(
    wb: Workbook, label: str, df: pd.DataFrame,
) -> str:
    """Write a pre-fetched World Bank DataFrame into a new worksheet."""
    prefix     = "WB - "
    safe_title = (prefix + label)[:31]
    existing   = {s.title for s in wb.worksheets}
    if safe_title in existing:
        safe_title = safe_title[:28] + "_2"

    ws = wb.create_sheet(safe_title)

    if df.empty:
        ws.cell(1, 1, "No data returned by the World Bank API for these parameters.")
        return safe_title

    write_headers(ws, ["Country", "Country Code", "Year", label])
    for i, row in df.iterrows():
        r = i + 2
        ws.cell(r, 1, row["Country"])
        ws.cell(r, 2, row["Country Code"])
        ws.cell(r, 3, int(row["Year"]))
        val_cell = ws.cell(r, 4, row["Value"])
        val_cell.number_format = '#,##0'
        style_row(ws, r, 4, i % 2 == 0)
    set_widths(ws, {"A": 28, "B": 14, "C": 8, "D": 26})
    return safe_title


def build_wb_indicator_sheet(
    wb: Workbook, label: str, indicator_code: str,
    country_codes: list, start_year: int, end_year: int,
) -> str:
    df = fetch_wb_indicator(indicator_code, tuple(country_codes), start_year, end_year)
    return _build_wb_sheet_from_df(wb, label, df)


# ── SOURCES SHEET ─────────────────────────────────────────────────────────────
def build_sources_sheet(wb: Workbook, registry: list[dict]):
    ws = wb.create_sheet("Sources", 0)

    title_cell = ws.cell(1, 1, "Data Sources")
    title_cell.font  = Font(bold=True, size=13, color="FFFFFF")
    title_cell.fill  = PatternFill("solid", fgColor="003366")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24
    ws.merge_cells("A1:E1")

    sub = ws.cell(2, 1, "Copy the sheet name + URL into your footnote / bibliography.")
    sub.font = Font(italic=True, size=9, color="595959")
    ws.merge_cells("A2:E2")

    write_headers(ws, ["Sheet Name", "Dataset", "Source Organisation", "URL", "Notes"], row=3)
    set_widths(ws, {"A": 26, "B": 38, "C": 30, "D": 54, "E": 42})

    link_font = Font(color="1155CC", underline="single")
    for i, entry in enumerate(registry):
        r = i + 4
        ws.cell(r, 1, entry["sheet_name"])
        ws.cell(r, 2, entry["dataset"])
        ws.cell(r, 3, entry["source_name"])
        uc = ws.cell(r, 4, entry["source_url"])
        uc.font = link_font
        ws.cell(r, 5, entry.get("notes", ""))
        style_row(ws, r, 5, i % 2 == 0)


# ── STATA DO-FILE GENERATOR ───────────────────────────────────────────────────
def _stata_varname(label: str) -> str:
    import re
    s = label.lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_")[:32]


def _stata_safe_dta(name: str) -> str:
    import re
    return re.sub(r"[^a-zA-Z0-9_\-]", "_", name).strip("_")


def generate_stata_do(
    sheet_registry: list[dict],
    excel_filename: str,
    today_str: str,
    fx_pair: tuple | None,
    fx_frequencies: list[str],
) -> str:
    wb_entries = [e for e in sheet_registry if e["sheet_name"].startswith("WB - ")]

    L: list[str] = []
    sep  = "=" * 74
    dash = "-" * 74

    # Every FX .dta file this do-file will produce, in order (daily is always first).
    fx_dta_files: list[str] = []
    if fx_pair:
        fx_dta_files.append("ecb_fx_daily.dta")
        fx_dta_files += [f"ecb_fx_{freq.lower()}.dta" for freq in fx_frequencies]

    # ── Header ───────────────────────────────────────────────────────────────
    L += [
        f"/* {sep}",
        f"   Import - Ancillary data - {today_str}",
        f"   Auto-generated by Ancillary data downloader on {today_str}",
        f"   Excel source : {excel_filename}",
        f"",
        f"   INSTRUCTIONS",
        f"   ------------",
        f"   1. Place this .do file in the SAME FOLDER as the Excel file.",
        f"   2. Set the global wd path below and uncomment the two lines.",
        f"   3. Run the script. Each block saves a .dta in the same folder.",
        f"",
        f"   OUTPUT FILES",
        f"   ------------",
    ]
    for f in fx_dta_files:
        L.append(f"   {f}")
    for e in wb_entries:
        L.append(f"   {_stata_safe_dta(e['sheet_name'])}.dta")
    if len(wb_entries) > 1:
        L.append(f"   WB_panel_combined.dta   (all WB indicators merged on countrycode + year)")
    L += [f"{sep} */", ""]

    L += [
        "* ── Working directory (edit and uncomment) ──────────────────────────── *",
        '* global wd "C:/your/folder/path"',
        '* cd "${wd}"',
        "",
        f'local excel "{excel_filename}"',
        "",
    ]

    # ── Section 1: FX ────────────────────────────────────────────────────────
    if fx_pair:
        base, quote = fx_pair
        fwd_col = f"{base}/{quote}"
        inv_col = f"{quote}/{base}"
        fwd_var = f"{base.lower()}_{quote.lower()}"   # e.g. usd_eur
        inv_var = f"{quote.lower()}_{base.lower()}"   # e.g. eur_usd

        L += [
            f"/* {dash}",
            f"   SECTION 1 — ECB Exchange Rates",
            f"   Pair   : {fwd_col}  (and inverse {inv_col})",
            f"   Source : {ECB_SOURCE['source_name']}",
            f"   URL    : {ECB_SOURCE['source_url']}",
            f"   Method : the Excel sheet holds only the daily (business-day) series.",
            f"            Every other frequency below is a simple average of that",
            f"            daily rate over the period — computed here, in full view,",
            f"            rather than pre-baked into the spreadsheet.",
            f"{dash} */",
            "",
            "* ---- Daily (base series, imported from Excel) ----",
            f'import excel using "`excel\'", sheet("ECB - FX - Daily") cellrange(A2) clear',
            "rename A date_str",
            f"rename B {fwd_var}",
            f"rename C {inv_var}",
            'gen date_daily = date(date_str, "YMD")',
            "format date_daily %td",
            "drop date_str",
            'label variable date_daily "Date"',
            f'label variable {fwd_var} "{fwd_col} — ECB reference rate"',
            f'label variable {inv_var} "{inv_col} — ECB reference rate"',
            f"order date_daily {fwd_var} {inv_var}",
            "sort date_daily",
            "tsset date_daily",
            'save "ecb_fx_daily.dta", replace',
            "",
        ]

        if "Weekly" in fx_frequencies:
            L += [
                "* ---- Weekly (Monday-start business week, average of the daily rate) ----",
                'use "ecb_fx_daily.dta", clear',
                "gen dow = dow(date_daily)                      // 0=Sun ... 6=Sat",
                "gen week_start = date_daily - mod(dow + 6, 7)   // Monday of that week",
                "format week_start %td",
                "drop dow",
                f"collapse (mean) {fwd_var} {inv_var}, by(week_start)",
                "gen week_end = week_start + 4",
                "format week_end %td",
                'label variable week_start "Week start (Monday)"',
                'label variable week_end   "Week end (Friday)"',
                f'label variable {fwd_var} "{fwd_col} — average of daily rates in the week"',
                f'label variable {inv_var} "{inv_col} — average of daily rates in the week"',
                f"order week_start week_end {fwd_var} {inv_var}",
                "tsset week_start",
                'save "ecb_fx_weekly.dta", replace',
                "",
            ]

        if "Monthly" in fx_frequencies:
            L += [
                "* ---- Monthly (average of the daily rate) ----",
                'use "ecb_fx_daily.dta", clear',
                "gen month = mofd(date_daily)",
                "format month %tm",
                f"collapse (mean) {fwd_var} {inv_var}, by(month)",
                'label variable month "Month"',
                f'label variable {fwd_var} "{fwd_col} — average of daily rates in the month"',
                f'label variable {inv_var} "{inv_col} — average of daily rates in the month"',
                "tsset month",
                'save "ecb_fx_monthly.dta", replace',
                "",
            ]

        if "Quarterly" in fx_frequencies:
            L += [
                "* ---- Quarterly (average of the daily rate) ----",
                'use "ecb_fx_daily.dta", clear',
                "gen quarter = qofd(date_daily)",
                "format quarter %tq",
                f"collapse (mean) {fwd_var} {inv_var}, by(quarter)",
                'label variable quarter "Quarter"',
                f'label variable {fwd_var} "{fwd_col} — average of daily rates in the quarter"',
                f'label variable {inv_var} "{inv_col} — average of daily rates in the quarter"',
                "tsset quarter",
                'save "ecb_fx_quarterly.dta", replace',
                "",
            ]

        if "Annual" in fx_frequencies:
            L += [
                "* ---- Annual (average of the daily rate) ----",
                'use "ecb_fx_daily.dta", clear',
                "gen year = yofd(date_daily)",
                "format year %ty",
                f"collapse (mean) {fwd_var} {inv_var}, by(year)",
                'label variable year "Year"',
                f'label variable {fwd_var} "{fwd_col} — average of daily rates in the year"',
                f'label variable {inv_var} "{inv_col} — average of daily rates in the year"',
                "tsset year",
                'save "ecb_fx_annual.dta", replace',
                "",
            ]

    # ── Section 2: World Bank ─────────────────────────────────────────────────
    if wb_entries:
        L += [
            f"/* {dash}",
            f"   SECTION 2 — World Bank Indicators",
            f"   Source : World Bank Open Data — World Development Indicators",
            f"   URL    : https://data.worldbank.org/indicator/",
            f"   Notes  : Annual frequency. Panel: countrycode (ISO-3) × year.",
            f"{dash} */",
            "",
        ]

        saved_dtas: list[tuple[str, str]] = []

        for entry in wb_entries:
            sname   = entry["sheet_name"]
            label   = entry["dataset"]
            code    = entry["source_url"].split("/")[-1]
            varname = _stata_varname(label)
            dta     = _stata_safe_dta(sname) + ".dta"
            note    = WB_INDICATOR_NOTES.get(code, "World Development Indicators (WDI).")

            L.append(f"* ── {label}  [{code}] {'─' * max(1, 50 - len(label))}")
            L.append(f'import excel using "`excel\'", sheet("{sname}") cellrange(A2) clear')
            L.append("")
            L.append("* Rename columns")
            L += ["rename A country", "rename B countrycode", "rename C year", f"rename D {varname}"]
            L.append("")
            L.append("* Variable labels")
            L += [
                'label variable country     "Country name"',
                'label variable countrycode "ISO-3 country code"',
                'label variable year        "Year"',
                f'label variable {varname} "{label} — {note}"',
            ]
            L.append("")
            L.append("destring year, replace")
            L.append("")
            L.append(f'save "{dta}", replace')
            L.append("")
            saved_dtas.append((dta, varname))

        if len(saved_dtas) > 1:
            L += [
                f"/* {dash}",
                f"   SECTION 2b — World Bank Combined Panel",
                f"   Merges all WB .dta files on countrycode + year.",
                f"{dash} */",
                "",
            ]
            L.append(f'use "{saved_dtas[0][0]}", clear')
            L.append("")
            for dta, var in saved_dtas[1:]:
                L += [f'merge 1:1 countrycode year using "{dta}", ///', f"    keepusing({var}) nogenerate", ""]
            L += [
                "encode countrycode, gen(country_id)",
                "xtset country_id year",
                "sort countrycode year",
                "",
                'save "WB_panel_combined.dta", replace',
                "",
            ]
        elif len(saved_dtas) == 1:
            L += ["* Only one WB indicator — no panel merge needed.", f'* Dataset saved as: "{saved_dtas[0][0]}"', ""]

    L += [f"/* {sep}", f"   End of do-file — Import - Ancillary data - {today_str}", f"{sep} */"]
    return "\n".join(L)


# ═════════════════════════════════════════════════════════════════════════════
# UI
# ═════════════════════════════════════════════════════════════════════════════
st.markdown("""
<div class="app-header">
  <h1>🌐 Ancillary data downloader</h1>
  <p>Select only the datasets you need — data is fetched on demand and exported to a single Excel file.</p>
</div>
""", unsafe_allow_html=True)


# ── SECTION 1 — FX ───────────────────────────────────────────────────────────
with st.expander("💱  Exchange Rates  (ECB)", expanded=True):
    include_fx = st.checkbox("Include exchange rate data", value=True)

    fx_base  = "USD"
    fx_quote = "EUR"

    if include_fx:
        c1, c2 = st.columns(2)
        with c1:
            fx_start = st.date_input("Start date", pd.to_datetime("2015-01-01"), key="fx_start")
        with c2:
            fx_end = st.date_input("End date", pd.to_datetime("today"), key="fx_end")
        if fx_start > fx_end:
            st.error("Start date must be before end date.")

        st.markdown("**Select your currency pair**")
        p1, p2 = st.columns(2)
        with p1:
            base_raw = st.selectbox(
                "🏦 I have (base currency)",
                options=CURRENCY_OPTIONS,
                index=CURRENCY_OPTIONS.index("USD — US Dollar"),
                key="fx_base",
                help="The currency you are converting FROM. E.g. if you want to know how many EUR one USD buys, select USD here.",
            )
        with p2:
            quote_raw = st.selectbox(
                "💰 I want (quote currency)",
                options=CURRENCY_OPTIONS,
                index=CURRENCY_OPTIONS.index("EUR — Euro"),
                key="fx_quote",
                help="The currency you are converting TO. The rate tells you how many units of this currency one unit of the base buys.",
            )

        fx_base  = base_raw.split(" — ")[0]
        fx_quote = quote_raw.split(" — ")[0]

        if fx_base == fx_quote:
            st.error("Base and quote currencies must be different.")
        else:
            fwd = f"{fx_base}/{fx_quote}"
            inv = f"{fx_quote}/{fx_base}"
            base_name  = CURRENCY_CATALOGUE[fx_base]
            quote_name = CURRENCY_CATALOGUE[fx_quote]
            st.markdown(
                f'<div class="info-banner">'
                f'✔ <strong>{fwd}</strong> — how many <strong>{fx_quote} ({quote_name})</strong> '
                f'you get for 1 <strong>{fx_base} ({base_name})</strong>'
                f'<br>The inverse <strong>{inv}</strong> will also be included in every sheet.'
                f'<br><small>Source: ECB Statistical Data Warehouse. Cross rates derived via EUR where needed.</small>'
                f'</div>',
                unsafe_allow_html=True,
            )
            st.caption("Produces 4 sheets: Annual · Monthly · Weekly (Mon–Fri) · Daily")


# ── SECTION 2 — World Bank ────────────────────────────────────────────────────
with st.expander("🏦  World Bank Indicators", expanded=False):
    include_wb = st.checkbox("Include World Bank data", value=False)

    selected_indicators: list[tuple[str, str]] = []
    selected_countries:  list[str]             = ["all"]
    wb_start_year = 2000
    wb_end_year   = 2023

    if include_wb:
        st.markdown("**Select indicators** — one Excel sheet per indicator")
        st.markdown(
            '<div class="info-banner">20 indicators across 6 themes.</div>',
            unsafe_allow_html=True,
        )

        for group_name, group_dict in WB_INDICATOR_GROUPS.items():
            chosen = st.multiselect(
                group_name,
                options=list(group_dict.keys()),
                default=[],
                key=f"grp_{group_name}",
                placeholder="— none selected —",
            )
            for label in chosen:
                selected_indicators.append((label, group_dict[label]))

        if selected_indicators:
            n = len(selected_indicators)
            names = ", ".join(f"<em>{l}</em>" for l, _ in selected_indicators)
            st.markdown(
                f'<div class="info-banner">✔ <strong>{n} indicator{"s" if n > 1 else ""} selected:</strong> {names}</div>',
                unsafe_allow_html=True,
            )

        st.markdown("**Year range**")
        y1, y2 = st.columns(2)
        with y1:
            wb_start_year = st.number_input("From year", min_value=1960, max_value=2024, value=2000, step=1)
        with y2:
            wb_end_year   = st.number_input("To year",   min_value=1960, max_value=2024, value=2023, step=1)

        st.markdown("**Countries**")

        COUNTRY_GROUPS = {
            "G7": ["Canada", "France", "Germany", "Italy", "Japan",
                   "United Kingdom", "United States"],
            "G20": ["Argentina", "Australia", "Brazil", "Canada", "China",
                    "France", "Germany", "India", "Indonesia", "Italy",
                    "Japan", "Korea, Rep.", "Mexico", "Russian Federation",
                    "Saudi Arabia", "South Africa", "Turkiye",
                    "United Kingdom", "United States"],
            "BRICS": ["Brazil", "Russian Federation", "India", "China",
                      "South Africa", "Egypt, Arab Rep.", "Ethiopia",
                      "Iran, Islamic Rep.", "Saudi Arabia",
                      "United Arab Emirates"],
            "EU": ["Austria", "Belgium", "Bulgaria", "Croatia", "Cyprus",
                   "Czechia", "Denmark", "Estonia", "Finland", "France",
                   "Germany", "Greece", "Hungary", "Ireland", "Italy",
                   "Latvia", "Lithuania", "Luxembourg", "Malta",
                   "Netherlands", "Poland", "Portugal", "Romania",
                   "Slovak Republic", "Slovenia", "Spain", "Sweden"],
            "OECD": ["Australia", "Austria", "Belgium", "Canada", "Chile",
                     "Colombia", "Costa Rica", "Czechia", "Denmark",
                     "Estonia", "Finland", "France", "Germany", "Greece",
                     "Hungary", "Iceland", "Ireland", "Israel", "Italy",
                     "Japan", "Korea, Rep.", "Latvia", "Lithuania",
                     "Luxembourg", "Mexico", "Netherlands", "New Zealand",
                     "Norway", "Poland", "Portugal", "Slovak Republic",
                     "Slovenia", "Spain", "Sweden", "Switzerland",
                     "Turkiye", "United Kingdom", "United States"],
            "EEA": ["Austria", "Belgium", "Bulgaria", "Croatia", "Cyprus", "Czechia",
                    "Denmark", "Estonia", "Finland", "France", "Germany", "Greece",
                    "Hungary", "Iceland", "Ireland", "Italy", "Latvia", "Liechtenstein",
                    "Lithuania", "Luxembourg", "Malta", "Netherlands", "Norway",
                    "Poland", "Portugal", "Romania", "Slovak Republic", "Slovenia",
                    "Spain", "Sweden"
            ]
        }

        def update_countries(country_list):
            st.session_state["wb_countries"] = country_list

        btn_cols = st.columns(7)
        for col, (group_name, group_list) in zip(btn_cols, COUNTRY_GROUPS.items()):
            col.button(group_name, on_click=update_countries, args=(group_list,),
                       key=f"btn_{group_name}", use_container_width=True)
        btn_cols[6].button("Clear", on_click=update_countries, args=([],),
                           key="btn_clear", use_container_width=True)

        try:
            all_countries_df = get_wb_countries()
            iso2_map         = dict(zip(all_countries_df["name"], all_countries_df["iso2"]))
            chosen_names     = st.multiselect(
                "Filter by country (optional)",
                options=all_countries_df["name"].tolist(),
                default=[],
                placeholder="All countries  (slower for long date ranges)",
                key="wb_countries",
            )
            if chosen_names:
                selected_countries = [iso2_map[n] for n in chosen_names]
            else:
                selected_countries = all_countries_df["iso2"].tolist()
            if not chosen_names:
                st.markdown(
                    '<div class="warn-banner">⚠ Be sure to select at least one country or region of interest.</div>',
                    unsafe_allow_html=True,
                )
        except Exception:
            st.warning("Could not load country list — all countries will be fetched.")
            selected_countries = ["all"]


# ── GENERATE ──────────────────────────────────────────────────────────────────
fx_ready = include_fx and (fx_base != fx_quote)
nothing_selected = (not fx_ready) and (not include_wb or not selected_indicators)

if nothing_selected:
    st.info("☝ Expand a section above and select at least one dataset to enable the download.")
else:
    if st.button("🚀  Generate & Download", type="primary"):
        all_ok = True

        if include_fx:
            if fx_start > fx_end:
                st.error("Fix the FX date range first.")
                all_ok = False
            if fx_base == fx_quote:
                st.error("Base and quote currencies must be different.")
                all_ok = False
        if include_wb and wb_start_year > wb_end_year:
            st.error("'From year' must be ≤ 'To year'.")
            all_ok = False

        if all_ok:
            wb_excel: Workbook         = Workbook()
            sheet_registry: list[dict] = []
            first_added                = False
            total_steps = (4 if fx_ready else 0) + len(selected_indicators)
            progress    = st.progress(0, text="Starting…")
            step        = 0

            # ── FX ────────────────────────────────────────────────────────
            if fx_ready:
                fwd_label = f"{fx_base}/{fx_quote}"
                progress.progress(0, text=f"Fetching {fwd_label} data from ECB…")
                try:
                    build_fx_sheets(wb_excel, fx_base, fx_quote, fx_start, fx_end)
                    first_added = True
                    for period in ["Daily", "Weekly", "Monthly", "Annual"]:
                        sheet_registry.append({
                            "sheet_name":  f"ECB - FX - {period}",
                            "dataset":     f"{fwd_label} Exchange Rate — {period}",
                            **ECB_SOURCE,
                        })
                    step += 4
                    progress.progress(step / total_steps, text=f"{fwd_label} data loaded ✓")
                except Exception as e:
                    st.error(f"Error fetching FX data: {e}")
                    all_ok = False

            # ── World Bank (concurrent fetch) ────────────────────────────
            if include_wb and all_ok:
                if not first_added:
                    wb_excel.active.title = "_tmp"
                    first_added = True

                progress.progress(step / total_steps, text="Fetching World Bank indicators…")
                with ThreadPoolExecutor(max_workers=100) as executor:
                    futures = {
                        executor.submit(
                            fetch_wb_indicator,
                            code,
                            tuple(selected_countries),
                            int(wb_start_year),
                            int(wb_end_year),
                        ): (label, code)
                        for label, code in selected_indicators
                    }

                    for future in as_completed(futures):
                        label, code = futures[future]
                        step += 1
                        progress.progress(step / total_steps, text=f"Processing: {label}…")
                        try:
                            df = future.result()
                            sname = _build_wb_sheet_from_df(wb_excel, label, df)
                            sheet_registry.append({
                                "sheet_name":  sname,
                                "dataset":     label,
                                "source_name": "World Bank Open Data",
                                "source_url":  f"https://data.worldbank.org/indicator/{code}",
                                "notes":       WB_INDICATOR_NOTES.get(code, "World Development Indicators (WDI)."),
                            })
                        except Exception as e:
                            st.warning(f"Could not fetch '{label}': {e}")

            if "_tmp" in wb_excel.sheetnames:
                del wb_excel["_tmp"]

            # ── Sources + save ────────────────────────────────────────────
            progress.progress(0.97, text="Writing Sources sheet…")
            build_sources_sheet(wb_excel, sheet_registry)
            progress.progress(1.0, text="Saving…")

            if all_ok:
                today_str = pd.Timestamp.today().strftime("%Y-%m-%d")

                tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
                wb_excel.save(tmp.name)
                with open(tmp.name, "rb") as f:
                    excel_bytes = f.read()

                fname_parts = []
                if fx_ready:
                    fname_parts.append(f"FX_{fx_base}_{fx_quote}_{fx_start}_to_{fx_end}")
                if include_wb and selected_indicators:
                    fname_parts.append(f"WB_{int(wb_start_year)}-{int(wb_end_year)}")
                excel_filename = "_".join(fname_parts) + ".xlsx"

                do_content  = generate_stata_do(sheet_registry, excel_filename, today_str, [(fx_base, fx_quote)] if fx_ready else [])
                do_filename = f"Import - Ancillary data - {today_str}.do"

                zip_buffer = BytesIO()
                with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
                    zf.writestr(excel_filename, excel_bytes)
                    zf.writestr(do_filename, do_content.encode("utf-8"))
                zip_buffer.seek(0)

                progress.empty()
                n = len(wb_excel.sheetnames)
                st.success(
                    f"✅ Done! {n} sheet{'s' if n != 1 else ''} in the Excel "
                    f"+ Stata .do file — both packed in the zip below."
                )

                st.download_button(
                    "📦  Download Excel + Stata .do  (.zip)",
                    zip_buffer,
                    file_name=f"Ancillary_data_{today_str}.zip",
                    mime="application/zip",
                )
